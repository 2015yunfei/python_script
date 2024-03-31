import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime


def replace_invalid_characters(name):
    invalid_characters = ['\\', '/', '*', '?', '[', ']', ':']
    for char in invalid_characters:
        name = name.replace(char, '_')
    return name


def get_folder_info(directory):
    folders_info = {}
    direct_files_info = []
    for root, dirs, files in os.walk(directory):
        if root != directory:
            folder_name = replace_invalid_characters(os.path.relpath(root, directory))
            folders_info[folder_name] = []
            for file in files:
                file_path = os.path.join(root, file)
                file_name, file_extension = os.path.splitext(file)
                file_size_bytes = os.path.getsize(file_path)
                file_size_mb = round(file_size_bytes / (1024 * 1024),
                                     1)  # Convert file size to MB and round to one decimal place
                modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                folders_info[folder_name].append((file_name, file_extension, file_size_mb, modified_time))
        else:
            for file in files:
                file_path = os.path.join(root, file)
                file_name, file_extension = os.path.splitext(file)
                file_size_bytes = os.path.getsize(file_path)
                file_size_mb = round(file_size_bytes / (1024 * 1024),
                                     1)  # Convert file size to MB and round to one decimal place
                modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                direct_files_info.append((file_name, file_extension, file_size_mb, modified_time))
    return folders_info, direct_files_info


def create_excel(folders_info, direct_files_info, excel_name):
    wb = Workbook()
    for folder, files_info in folders_info.items():
        ws = wb.create_sheet(title=folder)  # Create a new sheet for each folder
        ws.append(["文件名", "文件类型", "文件大小 (MB)", "修改时间"])
        for file_info in files_info:
            ws.append(file_info)
        # Set alignment for cells
        for col in range(2, ws.max_column + 1):
            for cell in ws[chr(64 + col)][1:]:
                cell.alignment = Alignment(horizontal='right')

    # Create a sheet for files directly under the current directory (not in subfolders)
    ws = wb.create_sheet(title="根目录文件")
    ws.append(["文件名", "文件类型", "文件大小 (MB)", "修改时间"])
    for file_info in direct_files_info:
        ws.append(file_info)
    # Set alignment for cells
    for col in range(2, ws.max_column + 1):
        for cell in ws[chr(64 + col)][1:]:
            cell.alignment = Alignment(horizontal='right')

    del wb["Sheet"]  # Delete the default sheet
    wb.save(excel_name)


if __name__ == "__main__":
    directory = os.getcwd()  # Get the current working directory
    folders_info, direct_files_info = get_folder_info(directory)
    create_excel(folders_info, direct_files_info, "文件信息.xlsx")
    print("Excel文件已生成！")
