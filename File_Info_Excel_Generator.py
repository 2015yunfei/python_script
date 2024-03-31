import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime


def get_file_info(directory):
    files_info = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            file_name, file_extension = os.path.splitext(file)
            file_size_bytes = os.path.getsize(file_path)
            file_size_mb = round(file_size_bytes / (1024 * 1024),
                                 1)  # Convert file size to MB and round to one decimal place
            created_time = datetime.fromtimestamp(os.path.getctime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
            modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
            files_info.append((file_name, file_extension, file_size_mb, created_time, modified_time, file_path))
    return files_info


def create_excel(files_info, excel_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["文件名", "文件类型", "文件大小 (MB)", "创建时间", "修改时间", "文件路径"])
    for file_info in files_info:
        ws.append(file_info)
    # Set alignment for cells
    for col in range(2, ws.max_column + 1):
        for cell in ws[chr(64 + col)][1:]:
            if col != 6:  # Skip left align for "文件路径" column
                cell.alignment = Alignment(horizontal='right')

    # Set column width
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add some padding and multiply by 1.2 for better appearance
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    wb.save(excel_name)


if __name__ == "__main__":
    directory = os.getcwd()  # Get the current working directory
    files_info = get_file_info(directory)
    create_excel(files_info, "文件信息.xlsx")
    print("Excel文件已生成！")
