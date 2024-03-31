import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime


def get_file_info(directory):
    files_info = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            file_name, file_extension = os.path.splitext(file)
            file_size_bytes = os.path.getsize(file_path)
            file_size_mb = round(file_size_bytes / (1024 * 1024), 1)  # 将文件大小转换为MB，并保留一位小数
            modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
            files_info.append((file_name, file_extension, file_size_mb, modified_time))
    return files_info


def create_excel(files_info, excel_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["文件名", "文件类型", "文件大小 (MB)", "修改时间"])
    for file_info in files_info:
        ws.append(file_info)
    # 设置除了文件名列之外的其他列为右对齐
    for col in range(2, ws.max_column + 1):
        for cell in ws[chr(64 + col)][1:]:
            cell.alignment = Alignment(horizontal='right')
    wb.save(excel_name)


if __name__ == "__main__":
    directory = os.getcwd()  # 获取当前工作目录
    files_info = get_file_info(directory)
    create_excel(files_info, "文件信息.xlsx")
    print("Excel文件已生成！")
